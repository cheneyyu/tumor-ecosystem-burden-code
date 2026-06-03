#!/usr/bin/env python3
"""Build submission-ready supplementary tables for the tumor ecosystem paper."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PANCAN = ROOT / "data" / "processed" / "pancancer_ecosystem"
VALIDATION = ROOT / "data" / "processed" / "translational_validation"
FIG = ROOT / "figures" / "translational_validation"
REPORTS = ROOT / "reports"
MS = ROOT / "manuscript"
OUT = ROOT / "submission" / "supplementary_tables"


@dataclass(frozen=True)
class SupplementTable:
    number: str
    sheet: str
    title: str
    source: Path | list[Path]
    description: str
    columns: list[str] | None = None
    sort_by: list[str] | None = None
    ascending: bool | list[bool] = True


TABLES = [
    SupplementTable(
        "S1",
        "S1_Data_Provenance",
        "External data provenance, availability, and checksum record",
        VALIDATION / "external_validation_provenance.tsv",
        "Public validation cohorts and raw-file provenance used for the manuscript.",
    ),
    SupplementTable(
        "S2",
        "S2_State_Signatures",
        "Standardized NMF ecosystem-state signature catalog",
        VALIDATION / "signature_catalog.tsv",
        "Marker genes used to score ecosystem states in TCGA, CPTAC, and ICB cohorts.",
        columns=[
            "state_id",
            "state_label",
            "compartment",
            "gene_symbol",
            "marker_rank_or_weight",
            "marker_delta",
            "state_mean",
            "rest_mean",
            "source_file",
        ],
        sort_by=["compartment", "state_id", "marker_rank_or_weight"],
    ),
    SupplementTable(
        "S3",
        "S3_Signature_Variants",
        "Signature variants for sensitivity analyses",
        VALIDATION / "signature_catalog_variants.tsv",
        "Top25, top50, positive-only, and cell-cycle-gene-excluded state signature variants.",
        sort_by=["state_id", "variant", "marker_rank_or_weight"],
    ),
    SupplementTable(
        "S4",
        "S4_State_Priority_Map",
        "Pan-cancer ecosystem-state progression, modeled cost, and prognosis map",
        PANCAN / "ecosystem_state_progression_cost_prognosis_map.csv",
        "Primary state-level map integrating modeled burden/cost scores, stage gradients, and TCGA prognosis.",
        sort_by=["rank_desc_cost"],
    ),
    SupplementTable(
        "S5",
        "S5_Cancer_State_Abundance",
        "Cancer-specific NMF ecosystem-state representation",
        PANCAN / "nmf_state_representation_by_cancer.csv",
        "Mean ecosystem-state representation by cancer type from the single-cell atlas.",
    ),
    SupplementTable(
        "S6",
        "S6_Stage_Gradients",
        "TCGA stage gradients for ecosystem states",
        PANCAN / "ecosystem_state_stage_gradients.csv",
        "Associations between ecosystem-state scores and tumor stage.",
        sort_by=["stage_fdr"],
    ),
    SupplementTable(
        "S7",
        "S7_Counterfactual",
        "Within-cancer counterfactual burden and modeled cost scores",
        PANCAN / "ecosystem_state_within_cancer_counterfactual_scores.csv",
        "Scenario-modeling estimates after separating within-cancer state variation from cancer-site composition.",
    ),
    SupplementTable(
        "S8",
        "S8_TCGA_Nested_Cox",
        "Nested TCGA Cox models for prioritized ecosystem states",
        VALIDATION / "tcga_incremental_cox.tsv",
        "Cancer-stratified Cox models testing incremental prognostic value beyond clinical and TME proxies.",
        sort_by=["model", "p_value"],
    ),
    SupplementTable(
        "S9",
        "S9_TCGA_Cancer_Meta",
        "Cancer-specific TCGA Cox fits and fixed-effect meta-analysis",
        VALIDATION / "tcga_cancer_specific_meta.tsv",
        "Cancer-specific survival effects and pan-cancer fixed-effect meta-analysis for priority states.",
        sort_by=["state_id", "row_type", "p_value"],
    ),
    SupplementTable(
        "S10",
        "S10_Robustness_Matrix",
        "Robustness evidence matrix for prioritized states",
        VALIDATION / "robustness_matrix.tsv",
        "Binary and semi-binary criteria summarizing confounding, sensitivity, counterfactual, and negative-control checks.",
        sort_by=["robustness_score"],
        ascending=False,
    ),
    SupplementTable(
        "S11",
        "S11_CPTAC_Clinical",
        "CPTAC UCEC clinical association tests",
        VALIDATION / "cptac_clinical_associations.tsv",
        "Open CPTAC UCEC RNA-derived state scores tested against stage, grade, overall survival, and recurrence-free survival.",
        sort_by=["scoring_method", "endpoint", "p_value"],
    ),
    SupplementTable(
        "S12",
        "S12_CPTAC_Protein",
        "CPTAC UCEC matched RNA/protein state and pathway correlations",
        VALIDATION / "cptac_protein_pathway_correlations.tsv",
        "Orthogonal proteomic anchoring of RNA-derived ecosystem-state scores.",
        sort_by=["comparison_type", "p_value"],
    ),
    SupplementTable(
        "S13",
        "S13_CPTAC_Coverage",
        "CPTAC signature and protein-pathway gene coverage",
        [
            VALIDATION / "cptac_signature_gene_coverage.tsv",
            VALIDATION / "cptac_protein_signature_gene_coverage.tsv",
            VALIDATION / "cptac_protein_pathway_gene_coverage.tsv",
        ],
        "Gene coverage for CPTAC RNA scores, state-level protein scores, and grouped protein pathway scores.",
    ),
    SupplementTable(
        "S14",
        "S14_ICB_Validation",
        "Public immunotherapy response and survival/PFS validation",
        VALIDATION / "icb_state_validation.tsv",
        "Riaz/GSE91061 melanoma nivolumab and IMvigor210 urothelial atezolizumab response and survival/PFS analyses.",
        sort_by=["cohort", "endpoint", "p_value"],
    ),
    SupplementTable(
        "S15",
        "S15_ICB_Coverage",
        "ICB cohort signature gene coverage",
        VALIDATION / "icb_signature_gene_coverage.tsv",
        "Gene coverage for ecosystem-state scoring in Riaz/GSE91061 and IMvigor210.",
    ),
    SupplementTable(
        "S16",
        "S16_Figure_Sources",
        "Source data for translational validation extension figures",
        [
            FIG / "tcga_incremental_value_source.tsv",
            FIG / "robustness_heatmap_source.tsv",
            FIG / "cptac_validation_source.tsv",
            FIG / "icb_validation_source.tsv",
        ],
        "Long-form source data for the extension figures.",
    ),
    SupplementTable(
        "S17",
        "S17_Key_Resources",
        "Key resources table",
        MS / "key_resources_table.tsv",
        "Datasets, software, scripts, and identifiers used in the translational validation extension.",
    ),
    SupplementTable(
        "S18",
        "S18_File_Manifest",
        "Generated file manifest with checksums",
        REPORTS / "translational_validation_file_manifest.tsv",
        "Generated output manifest recording file paths, descriptions, dependencies, and SHA256 checksums.",
    ),
]


def read_table(path: Path) -> pd.DataFrame:
    suffixes = "".join(path.suffixes)
    if suffixes.endswith(".csv.gz"):
        return pd.read_csv(path)
    if path.suffix == ".csv":
        return pd.read_csv(path)
    return pd.read_csv(path, sep="\t")


def safe_name(text: str) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return text[:80]


def load_submission_table(spec: SupplementTable) -> pd.DataFrame:
    sources = spec.source if isinstance(spec.source, list) else [spec.source]
    frames = []
    for path in sources:
        df = read_table(path)
        df.insert(0, "supplement_source_file", str(path.relative_to(ROOT)))
        frames.append(df)
    df = pd.concat(frames, ignore_index=True, sort=False)
    if spec.columns:
        keep = ["supplement_source_file"] + [c for c in spec.columns if c in df.columns]
        extra = [c for c in df.columns if c not in keep]
        df = df[keep + extra]
    if spec.sort_by:
        cols = [c for c in spec.sort_by if c in df.columns]
        if cols:
            df = df.sort_values(cols, ascending=spec.ascending, na_position="last")
    return df


def write_index(index_df: pd.DataFrame) -> None:
    index_df.to_csv(OUT / "Supplementary_Table_Index.tsv", sep="\t", index=False)
    lines = [
        "# Supplementary Tables README",
        "",
        "This directory contains a submission-ready supplementary table package.",
        "",
        "Primary upload file:",
        "",
        "- `Supplementary_Tables.xlsx`: multi-sheet workbook. Each sheet corresponds to one numbered supplementary table.",
        "",
        "Machine-readable backup files:",
        "",
        "- `Supplementary_Table_SXX_*.tsv`: one tab-delimited file per supplementary table.",
        "- `Supplementary_Table_Index.tsv`: table titles, descriptions, sources, and dimensions.",
        "",
        "Usage notes:",
        "",
        "- Cost and burden columns are modeled/ecological and should not be described as observed patient-level spending.",
        "- CPTAC validation is open-access UCEC only; it supports independent grade/protein anchoring, not pan-cancer clinical validation by itself.",
        "- ICB analyses are exploratory cohort-specific response and OS/PFS association tests.",
        "",
        "## Table Index",
        "",
        "| table | sheet | title | rows | columns |",
        "|---|---|---|---:|---:|",
    ]
    for _, r in index_df.iterrows():
        lines.append(f"| {r['table_number']} | {r['sheet_name']} | {r['title']} | {r['n_rows']} | {r['n_columns']} |")
    (OUT / "README_supplementary_tables.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            values = ["" if c.value is None else str(c.value) for c in column_cells[:200]]
            width = min(max(10, max((len(v) for v in values), default=10) + 2), 55)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 500)):
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical="top")
    wb.save(path)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    index_rows = []
    data_frames = {}
    for spec in TABLES:
        df = load_submission_table(spec)
        file_name = f"Supplementary_Table_{spec.number}_{safe_name(spec.sheet)}.tsv"
        df.to_csv(OUT / file_name, sep="\t", index=False)
        data_frames[spec.sheet] = df
        source_paths = spec.source if isinstance(spec.source, list) else [spec.source]
        index_rows.append(
            {
                "table_number": spec.number,
                "sheet_name": spec.sheet,
                "title": spec.title,
                "description": spec.description,
                "source_files": "; ".join(str(p.relative_to(ROOT)) for p in source_paths),
                "tsv_file": file_name,
                "n_rows": len(df),
                "n_columns": len(df.columns),
            }
        )

    index_df = pd.DataFrame(index_rows)
    write_index(index_df)
    workbook = OUT / "Supplementary_Tables.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        index_df.to_excel(writer, sheet_name="Table_Index", index=False)
        for sheet, df in data_frames.items():
            df.to_excel(writer, sheet_name=sheet[:31], index=False)
    style_workbook(workbook)
    print(f"Wrote {workbook}")
    print(f"Wrote {len(TABLES)} supplementary TSV tables to {OUT}")


if __name__ == "__main__":
    main()
